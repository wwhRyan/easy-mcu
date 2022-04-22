import sys
from openpyxl import load_workbook
path = '190604_CXD3554GG_register_setting_SXRD241_60Hz (E).xlsx'
sheet_name = 'MISC'
start_row = 5 - 1
end_row = 165

if(len(sys.argv) != 5):
    exit(1)
path = sys.argv[1]
sheet_name = sys.argv[2]
start_row = int(sys.argv[3])
end_row = int(sys.argv[4])

# data_only=True, read_only=True faster
wb = load_workbook(filename=path, data_only=True, read_only=True)

ws = wb[sheet_name]
data = tuple(ws)

# output head file header
print("#ifndef __" + sheet_name + "_H__")
print("#define __" + sheet_name + "_H__")
print("")
print("#include <stdint.h>")
print("")
print("#define {}_BASE_ADDRESS 0x{}".format(sheet_name, ws.cell(
    row=2, column=4).value.split(': ')[1].split('h')[0]))
print("#define {}_REG_NUM {}".format(sheet_name, end_row - start_row))
print("#define {}_REG_ADDR(reg) REG_ADDR({}_BASE_ADDRESS,{}_t,reg)".format(
    sheet_name, sheet_name, sheet_name.lower()))
print("")
# foreach row
for row in range(start_row, end_row):
    # for j in range(2,10):
    #     print(data[i][j].value)
    idx = 0
    line = {}
    # foreach column
    for col in range(2, 10):
        # get column string in line[8]
        line[idx] = data[row][col].value
        idx += 1

    # calculate line[8] none element bits in count[8]
    count = [1, 1, 1, 1, 1, 1, 1, 1]
    for idx in range(0, 8):
        if(line[idx] == None and count[idx] != 0):
            count[idx] = count[idx] - 1
            while line[idx] == None:
                idx -= 1
            count[idx] += 1

    # remove "-" elements in line[8]
    for idx in range(0, 8):
        if line[idx] == "-":
            line[idx] = " "
        # cut off the useless '[x:x]'
        if(line[idx] != None):
            line[idx] = line[idx].split("[")[0]

    # output line[8] element bits
    print("union {}_REG_{:02X} {{".format(sheet_name, row - 4))
    print("uint8_t byte;    struct {")

    # foreach line[8] element bits using little endian
    for idx in range(7, -1, -1):
        if(count[idx] != 0):
            print("uint8_t {0}:{1};".format(line[idx], count[idx]))
    print("};};")

print("typedef struct {0}_T {{".format(sheet_name))
for idx in range(0, end_row - start_row):
    print("union {}_REG_{:02X} reg{:02X};".format(sheet_name, idx, idx))
print("}} {:s}_t;".format(sheet_name.lower()))

print("")
print("extern {:s}_t {:s};".format(sheet_name.lower(), sheet_name.lower()))
print("")

print("#endif")
